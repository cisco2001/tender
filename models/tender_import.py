import base64
import io
import logging
from datetime import datetime, timedelta
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

try:
    import pandas as pd
    import openpyxl
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None
    openpyxl = None

_logger = logging.getLogger(__name__)


class TenderImport(models.Model):
    _name = 'tender.import'
    _description = 'Tender Import'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'

    name = fields.Char(string='Import Name', required=True, default='New Import')
    state = fields.Selection([
        ('preview', 'Preview Available'),
        ('imported', 'Imported'),
        ('error', 'Error')
    ], string='Status', default='preview', tracking=True)
    
    institution_id = fields.Many2one(
        'res.partner', 
        string='Institution',
        help='Institution to which the uploaded tenders belong'
    )
    
    # File Upload
    excel_file = fields.Binary(
        string='Excel File',
        help='Upload Excel file containing tender data'
    )
    filename = fields.Char(string='Filename')
    
    # Import Configuration with Contact Information
    institution_name = fields.Char(
        string='Institution Name',
        help='Name of the institution these tenders belong to'
    )
    contact_person = fields.Char(
        string='Contact Person',
        help='Main contact person at the institution'
    )
    contact_email = fields.Char(
        string='Contact Email',
        help='Email address for tender inquiries'
    )
    contact_phone = fields.Char(
        string='Contact Phone',
        help='Phone number for tender inquiries'
    )
    
    # Import Results
    total_rows = fields.Integer(string='Total Rows', readonly=True)
    parsed_count = fields.Integer(string='Parsed Count', readonly=True)
    imported_count = fields.Integer(string='Import Count', readonly=True)
    error_count = fields.Integer(string='Error Count', readonly=True)
    
    # Relations
    preview_line_ids = fields.One2many(
        'tender.import.line',
        'import_id',
        string='Preview Lines'
    )
    tender_ids = fields.One2many(
        'tender.tender',
        'import_id',
        string='Imported Tenders'
    )
    
    # Error Handling
    error_message = fields.Text(string='Error Message', readonly=True)
    import_log = fields.Text(string='Import Log', readonly=True)

    @api.model
    def create(self, vals):
        """Override create to set proper name and copy institution info"""
        if vals.get('name') == 'New Import':
            vals['name'] = f"Import {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # If institution_id is set, copy contact info from partner
        if vals.get('institution_id') and not vals.get('institution_name'):
            institution = self.env['res.partner'].browse(vals['institution_id'])
            vals.update({
                'institution_name': institution.name,
                'contact_email': institution.email or '',
                'contact_phone': institution.phone or institution.mobile or '',
                'contact_person': institution.name
            })
        
        return super().create(vals)

    @api.onchange('institution_id')
    def _onchange_institution_id(self):
        """Auto-fill institution contact info when partner is selected"""
        if self.institution_id:
            self.institution_name = self.institution_id.name
            self.contact_email = self.institution_id.email or ''
            self.contact_phone = self.institution_id.phone or self.institution_id.mobile or ''
            self.contact_person = self.institution_id.name

    @api.constrains('contact_email')
    def _check_contact_email(self):
        """Validate email format"""
        for record in self:
            if record.contact_email and '@' not in record.contact_email:
                raise ValidationError("Please enter a valid email address")

    @api.onchange('excel_file')
    def _onchange_excel_file(self):
        """Automatically process file when uploaded"""
        if self.excel_file:
            try:
                self.action_upload_file()
            except Exception as e:
                # Silently handle errors to prevent blocking the user
                _logger.error(f"Auto-process failed: {str(e)}", exc_info=True)

    def action_upload_file(self):
        """Process uploaded Excel file - simplified version without logs"""
        _logger.info("=== STARTING EXCEL UPLOAD PROCESS ===")
        
        # Clear previous state
        self.preview_line_ids.unlink()
        self.error_message = False
        self.import_log = False
        
        # Validate file upload
        if not self.excel_file:
            raise UserError("Please upload an Excel file first.")
        
        if not self.filename:
            raise UserError("Filename is required.")
        
        # Check pandas availability
        if not PANDAS_AVAILABLE:
            error_msg = "Required libraries not installed. Please install: pip install pandas openpyxl xlrd"
            self.error_message = error_msg
            self.state = 'error'
            _logger.error(error_msg)
            raise UserError(error_msg)
        
        try:
            # Process the file
            self._process_excel_file()
            
            # State is already preview by default
            _logger.info("Excel processing completed successfully")
            
        except Exception as e:
            self.state = 'error'
            error_msg = f"Error processing Excel file: {str(e)}"
            self.error_message = error_msg
            _logger.error(f"Excel processing error: {e}", exc_info=True)
            raise UserError(error_msg)

    def _process_excel_file(self):
        """Core Excel processing logic without logging to user"""
        _logger.info("Starting Excel file processing")
        
        try:
            # Decode the binary file
            file_data = base64.b64decode(self.excel_file)
            _logger.info(f"File decoded, size: {len(file_data)} bytes")
            
            # Read Excel file with multiple fallback methods
            df = self._read_excel_file(file_data)
            
            # Validate DataFrame
            if df is None or df.empty:
                raise UserError("Excel file is empty or could not be read")
            
            _logger.info(f"Excel file loaded: {df.shape[0]} rows, {df.shape[1]} columns")
            
            # Process the data
            self.total_rows = len(df)
            self._parse_excel_data(df)
            
        except Exception as e:
            _logger.error(f"Error in _process_excel_file: {e}", exc_info=True)
            raise

    def _read_excel_file(self, file_data):
        """Try multiple methods to read Excel file"""
        _logger.info("Attempting to read Excel file")
        
        # Method 1: Try with pandas read_excel (most common)
        try:
            _logger.info("Trying pandas read_excel...")
            df = pd.read_excel(io.BytesIO(file_data), sheet_name=0)
            if not df.empty:
                _logger.info("✅ Success with pandas read_excel")
                return df
        except Exception as e:
            _logger.warning(f"pandas read_excel failed: {e}")

        # Method 2: Try with openpyxl engine specifically
        try:
            _logger.info("Trying pandas with openpyxl engine...")
            df = pd.read_excel(io.BytesIO(file_data), sheet_name=0, engine='openpyxl')
            if not df.empty:
                _logger.info("✅ Success with openpyxl engine")
                return df
        except Exception as e:
            _logger.warning(f"openpyxl engine failed: {e}")

        # Method 3: Try with xlrd engine (for older Excel files)
        try:
            _logger.info("Trying pandas with xlrd engine...")
            df = pd.read_excel(io.BytesIO(file_data), sheet_name=0, engine='xlrd')
            if not df.empty:
                _logger.info("✅ Success with xlrd engine")
                return df
        except Exception as e:
            _logger.warning(f"xlrd engine failed: {e}")

        # Method 4: Try direct openpyxl approach
        try:
            _logger.info("Trying direct openpyxl...")
            from openpyxl import load_workbook
            
            workbook = load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            # Convert to pandas DataFrame
            data = []
            headers = []
            
            # Get headers from first row
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Get data from remaining rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row)
            
            df = pd.DataFrame(data, columns=headers)
            if not df.empty:
                _logger.info("✅ Success with direct openpyxl")
                return df
                
        except Exception as e:
            _logger.warning(f"Direct openpyxl failed: {e}")

        # If all methods fail
        raise UserError("Could not read Excel file. Please ensure it's a valid Excel file (.xlsx, .xls)")

    def _parse_excel_data(self, df):
        """Parse Excel DataFrame and create preview lines"""
        _logger.info("Starting data parsing")
        
        # Get column mappings
        column_mappings = self._get_column_mappings(df.columns)
        
        # Process each row
        preview_lines = []
        successful_rows = 0
        error_rows = 0
        skipped_rows = 0
        
        for index, row in df.iterrows():
            try:
                # Skip completely empty rows
                if row.isna().all():
                    skipped_rows += 1
                    continue
                
                # Extract data from row
                line_data = self._extract_row_data(row, column_mappings)
                
                # Validate extracted data
                if self._validate_line_data(line_data):
                    preview_lines.append((0, 0, line_data))
                    successful_rows += 1
                else:
                    error_rows += 1
                    _logger.warning(f"Invalid data in row {index + 1}")
                    
            except Exception as e:
                error_rows += 1
                _logger.error(f"Error processing row {index + 1}: {e}")
        
        # Create preview lines
        if preview_lines:
            self.preview_line_ids = preview_lines
        
        # Update counts
        self.parsed_count = successful_rows
        self.error_count = error_rows
        
        _logger.info(f"Parsing completed: {successful_rows} success, {error_rows} errors, {skipped_rows} skipped")

    def _get_column_mappings(self, columns):
        """Map Excel columns to tender fields"""
        _logger.info(f"Mapping columns: {list(columns)}")
        
        # Convert to lowercase for matching
        columns_lower = [str(col).lower().strip() for col in columns]
        mappings = {}
        
        # Define mapping patterns
        mapping_patterns = {
            'name': ['title', 'tender title', 'name', 'tender name', 'subject', 'tender', 'description'],
            'reference': ['reference', 'ref', 'tender ref', 'reference number', 'number', 'id', 'tender id'],
            'description': ['description', 'desc', 'details', 'summary', 'info', 'information'],
            'deadline_date': ['deadline', 'due date', 'closing date', 'submission date', 'end date', 'close'],
            'publication_date': ['publication date', 'published', 'announcement date', 'start date', 'publish'],
            'contact_email': ['email', 'contact email', 'e-mail', 'contact_email'],
            'contact_phone': ['phone', 'telephone', 'contact phone', 'mobile', 'contact_phone'],
        }
        
        # Find best matches
        for field, patterns in mapping_patterns.items():
            for i, col_lower in enumerate(columns_lower):
                if field not in mappings:  # Only map if not already mapped
                    for pattern in patterns:
                        if pattern in col_lower:
                            mappings[field] = columns[i]
                            _logger.info(f"Mapped '{field}' to column '{columns[i]}'")
                            break
        
        # Ensure we have at least a name field (fallback to first column)
        if 'name' not in mappings and columns:
            mappings['name'] = columns[0]
            _logger.info(f"Using first column '{columns[0]}' as name field")
        
        return mappings

    def _extract_row_data(self, row, mappings):
        """Extract tender data from a single row"""
        data = {
            'import_id': self.id,
            'institution': self.institution_name or '',
        }
        
        # Extract basic fields
        data['name'] = self._safe_get_value(row, mappings.get('name')) or 'Unnamed Tender'
        data['reference'] = self._safe_get_value(row, mappings.get('reference'))
        data['description'] = self._safe_get_value(row, mappings.get('description'))
        
        # Extract contact info from Excel if available, otherwise use import default
        excel_email = self._safe_get_value(row, mappings.get('contact_email'))
        excel_phone = self._safe_get_value(row, mappings.get('contact_phone'))
        
        data['contact_email'] = excel_email or self.contact_email or ''
        data['contact_phone'] = excel_phone or self.contact_phone or ''

        # Extract and parse dates
        deadline_str = self._safe_get_value(row, mappings.get('deadline_date'))
        if deadline_str:
            data['deadline_date'] = self._parse_date(deadline_str)
        
        pub_date_str = self._safe_get_value(row, mappings.get('publication_date'))
        if pub_date_str:
            data['publication_date'] = self._parse_date(pub_date_str)
        
        # Auto-categorize and set initial IT-related flag
        category_id = self._auto_categorize(data['name'], data.get('description', ''))
        if category_id:
            data['category_id'] = category_id
            # Check if the category is marked as IT-related
            category = self.env['tender.category'].browse(category_id)
            data['is_it_related'] = category.is_it_related
        
        # Check for IT keywords if not already marked as IT-related
        if 'is_it_related' not in data or not data['is_it_related']:
            search_text = f"{data['name']} {data.get('description', '')}".lower()
            it_keywords = [
                'software', 'hardware', 'network', 'computer', 'system',
                'application', 'database', 'server', 'cloud', 'cybersecurity',
                'programming', 'development', 'it services', 'technology'
            ]
            data['is_it_related'] = any(keyword in search_text for keyword in it_keywords)
        
        return data

    def _safe_get_value(self, row, column):
        """Safely extract value from row"""
        if not column or column not in row.index:
            return ''
        
        value = row[column]
        if pd.isna(value):
            return ''
        
        return str(value).strip()

    def _parse_date(self, date_str):
        """Parse date from various formats"""
        if not date_str:
            return False
        
        # Handle pandas Timestamp objects
        if hasattr(date_str, 'date'):
            return date_str.date()
        
        # Try to parse string dates
        date_formats = [
            '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', 
            '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y'
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(str(date_str).split()[0], fmt).date()
            except ValueError:
                continue
        
        return False

    def _auto_categorize(self, title, description):
        """Auto-categorize tender based on content"""
        try:
            return self.env['tender.tender'].auto_categorize_tender(title, description)
        except:
            return False

    def _validate_line_data(self, line_data):
        """Validate extracted line data"""
        # Must have a name
        if not line_data.get('name') or line_data['name'] == 'Unnamed Tender':
            return False
        
        # Name should be reasonably long
        if len(line_data['name']) < 3:
            return False
        
        return True

    def action_import_selected(self):
        """Import tenders marked as IT-related by the user"""
        # Get all lines marked as IT-related (either by system or manually by user)
        it_related_lines = self.preview_line_ids.filtered(lambda l: l.is_it_related)
        
        if not it_related_lines:
            raise UserError("No tenders marked as IT-related. Please mark tenders as IT-related to import them.")
        
        imported_count = 0
        error_count = 0
        error_messages = []
        
        # Use a savepoint to handle individual import errors
        for line in it_related_lines:
            try:
                # Create a savepoint for each tender import
                with self.env.cr.savepoint():
                    # Prepare tender data with proper validation
                    tender_data = {
                        'name': line.name or 'Unnamed Tender',
                        'reference': line.reference or '',
                        'description': line.description or '',
                        'institution': line.institution or self.institution_name or '',
                        'category_id': line.category_id.id if line.category_id else False,
                        'assigned_to': line.assigned_to.id if line.assigned_to else self.env.user.id,
                        'import_id': self.id,
                        'state': 'draft',
                        'is_it_related': True,  # Mark as IT-related in the tender record
                        'contact_email': line.contact_email or self.contact_email or '',
                        'contact_phone': line.contact_phone or self.contact_phone or '',
                        'contact_person': self.contact_person or '',
                    }
                    
                    # Handle deadline_date - set to future date if missing
                    if line.deadline_date:
                        tender_data['deadline_date'] = line.deadline_date
                    else:
                        # Set default deadline to 30 days from today if missing
                        from datetime import datetime, timedelta
                        tender_data['deadline_date'] = (datetime.now() + timedelta(days=30)).date()
                        _logger.warning(f"Missing deadline for tender '{line.name}', set to 30 days from today")
                    
                    # Handle publication_date - optional field
                    if line.publication_date:
                        tender_data['publication_date'] = line.publication_date
                    
                    # Check if tender.tender model exists and has required fields
                    if 'tender.tender' not in self.env:
                        raise UserError("Tender model (tender.tender) not found. Please ensure the tender module is properly installed.")
                    
                    # Create the tender record
                    tender = self.env['tender.tender'].create(tender_data)
                    imported_count += 1
                    _logger.info(f"Successfully imported IT tender: {line.name}")
                    
            except Exception as e:
                error_count += 1
                error_msg = f"Error importing '{line.name}': {str(e)}"
                error_messages.append(error_msg)
                _logger.error(error_msg, exc_info=True)
        
        # Update counters using a separate transaction
        try:
            with self.env.cr.savepoint():
                # Update the import record
                self.write({
                    'imported_count': imported_count,
                    'error_count': error_count,
                    'state': 'imported' if imported_count > 0 else 'error'
                })
                
                if error_messages:
                    self.error_message = '\n'.join(error_messages)
        
        except Exception as e:
            _logger.error(f"Error updating import record: {e}", exc_info=True)
            raise UserError(f"Import completed but failed to update record: {str(e)}")
        
        # Show appropriate message and action
        if imported_count > 0:
            message = f"Successfully imported {imported_count} IT tenders"
            if error_count > 0:
                message += f" ({error_count} errors)"
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Import Completed',
                    'message': message,
                    'type': 'success' if error_count == 0 else 'warning',
                    'sticky': False,
                },
                'context': {
                    'next_action': {
                        'type': 'ir.actions.act_window',
                        'name': 'Imported Tenders',
                        'res_model': 'tender.tender',
                        'view_mode': 'tree,form,kanban',
                        'domain': [('import_id', '=', self.id)],
                    }
                }
            }
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Import Failed',
                    'message': f"No IT tenders were imported. {error_count} errors occurred.",
                    'type': 'danger',
                    'sticky': True,
                }
            }
    
    def action_view_tenders(self):
        """View imported tenders"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Imported Tenders',
            'res_model': 'tender.tender',
            'view_mode': 'tree,form,kanban',
            'domain': [('import_id', '=', self.id)],
        }


class TenderImportLine(models.Model):
    _name = 'tender.import.line'
    _description = 'Tender Import Line'
    _order = 'sequence, id'

    import_id = fields.Many2one('tender.import', string='Import', required=True, ondelete='cascade')
    sequence = fields.Integer(string='Sequence', default=10)
    
    # Tender Data
    name = fields.Char(string='Tender Title', required=True)
    reference = fields.Char(string='Reference')
    description = fields.Text(string='Description')
    institution = fields.Char(string='Institution')
    
    # Dates
    deadline_date = fields.Date(string='Deadline Date')
    publication_date = fields.Date(string='Publication Date')
    
    # Contact Information (kept in model but hidden from preview)
    contact_email = fields.Char(string='Contact Email')
    contact_phone = fields.Char(string='Contact Phone')
    
    # Assignment
    category_id = fields.Many2one('tender.category', string='Category')
    assigned_to = fields.Many2one('res.users', string='Assign To')
    
    # IT-related flag - fully editable by user
    is_it_related = fields.Boolean(
        string='IT Related',
        help="Mark this tender as IT-related for import"
    )